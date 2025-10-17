"""Data loading utilities for the Tecnoloc reconciliation pipeline."""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Tuple

from .models import ErpRecord, PayfyExpense


class DataValidationError(RuntimeError):
    """Raised when the input file is invalid."""


def _parse_date(value: str, *, field_name: str) -> datetime:
    """Parse a textual date supporting the formats used in the reports."""

    if not value:
        raise DataValidationError(f"Data vazia no campo '{field_name}'.")

    normalized = value.strip()
    formats = (
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
    )
    for date_format in formats:
        try:
            return datetime.strptime(normalized, date_format)
        except ValueError:
            continue
    raise DataValidationError(f"Data inválida '{value}' no campo '{field_name}'.")


def _parse_float(value: str, *, field_name: str) -> float:
    normalized = value.replace("R$", "").replace("$", "").replace(" ", "").replace(".", "").replace(",", ".")
    try:
        return float(normalized)
    except ValueError as exc:
        raise DataValidationError(f"Valor inválido '{value}' no campo '{field_name}'.") from exc


def _read_csv_rows(path: Path) -> Iterator[Dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as csv_file:
        preview = csv_file.readline()
        csv_file.seek(0)
        delimiter = ";" if preview.count(";") > preview.count(",") else ","
        reader = csv.DictReader(csv_file, delimiter=delimiter)
        if reader.fieldnames is None:
            return iter(())
        for row in reader:
            cleaned = {
                (key or "").strip(): (value.strip() if isinstance(value, str) else "" if value is None else str(value))
                for key, value in row.items()
            }
            if not any(cleaned.values()):
                continue
            yield cleaned


def _read_excel_rows(path: Path) -> Iterator[Dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise DataValidationError(
            "Leitura de arquivos .xlsx requer a dependência 'openpyxl'."
        ) from exc

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            return

        normalized_headers = [str(header).strip() if header is not None else "" for header in headers]
        for values in rows:
            record = {}
            empty = True
            for header, value in zip(normalized_headers, values or ()):  # type: ignore[arg-type]
                if header == "":
                    continue
                if isinstance(value, datetime):
                    text = value.strftime("%d/%m/%Y %H:%M:%S")
                elif value is None:
                    text = ""
                else:
                    text = str(value).strip()
                if text:
                    empty = False
                record[header] = text
            if empty:
                continue
            yield record
    finally:
        workbook.close()


def _read_rows(path: Path) -> Iterable[Dict[str, str]]:
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"Arquivo não encontrado: {path}")

    suffix = path.suffix.lower()
    if suffix in {".csv"}:
        return list(_read_csv_rows(path))
    if suffix in {".xlsx", ".xlsm"}:
        return list(_read_excel_rows(path))
    raise DataValidationError(f"Formato de arquivo não suportado: '{path.suffix}'.")


def load_payfy_expenses(path: Path) -> List[PayfyExpense]:
    """Load the detailed PayFy expenses file."""

    required = {"Usuário", "Data Transação", "Valor", "Status da Nota", "Categoria", "ID"}
    expenses: List[PayfyExpense] = []
    for row in _read_rows(path):
        if not required.issubset(row):
            missing = ", ".join(sorted(required - set(row)))
            raise DataValidationError(f"Campos obrigatórios ausentes: {missing}.")
        transaction_date = _parse_date(row["Data Transação"], field_name="Data Transação")
        approval = row.get("Data da Aprovação")
        approval_dt = _parse_date(approval, field_name="Data da Aprovação") if approval else None
        value = _parse_float(row["Valor"], field_name="Valor")
        expense = PayfyExpense(
            user=row["Usuário"],
            date=transaction_date,
            value=value,
            status=row.get("Status da Nota", ""),
            category=row.get("Categoria", ""),
            expense_id=row.get("ID"),
            approval_date=approval_dt,
            raw=row,
        )
        expenses.append(expense)
    return expenses


def load_erp_records(path: Path) -> List[ErpRecord]:
    """Load the ERP balance report."""

    required = {
        "Data",
        "Usuário",
        "Carga Empresa",
        "Carga Cartão",
        "Descarga Cartão",
        "Tarifas",
        "Reembolsos",
        "Saldo Empresa",
    }
    records: List[ErpRecord] = []
    for row in _read_rows(path):
        if not required.issubset(row):
            missing = ", ".join(sorted(required - set(row)))
            raise DataValidationError(f"Campos obrigatórios ausentes: {missing}.")
        date = _parse_date(row["Data"], field_name="Data")
        value_fields = ["Carga Empresa", "Carga Cartão", "Descarga Cartão", "Tarifas", "Reembolsos"]
        for field in value_fields:
            amount = _parse_float(row[field], field_name=field)
            if amount == 0:
                continue
            entry_type = "Tarifa" if field == "Tarifas" else field
            records.append(
                ErpRecord(
                    user=row.get("Usuário", ""),
                    date=date,
                    value=amount,
                    erp_type=entry_type,
                    raw=row,
                )
            )
    return records


def load_payfy_card_summary(path: Path) -> List[Tuple[str, float, float]]:
    """Load the PayFy technicians card summary (team, initial/final balances)."""

    required = {"Time", "Saldo Inicial", "Saldo Final"}
    summary: List[Tuple[str, float, float]] = []
    for row in _read_rows(path):
        if not required.issubset(row):
            missing = ", ".join(sorted(required - set(row)))
            raise DataValidationError(f"Campos obrigatórios ausentes: {missing}.")
        team = row["Time"]
        initial_balance = _parse_float(row["Saldo Inicial"], field_name="Saldo Inicial")
        final_balance = _parse_float(row["Saldo Final"], field_name="Saldo Final")
        summary.append((team, initial_balance, final_balance))
    return summary


def load_erp_movements(path: Path) -> List[datetime]:
    """Load the ERP movement report to analyse the `data_mov` competence."""

    rows = list(_read_rows(path))
    if not rows:
        return []

    dates: List[datetime] = []
    for index, row in enumerate(rows, start=2):
        matching_key = next((key for key in row if key.lower() == "data_mov"), None)
        if not matching_key:
            raise DataValidationError("Campo obrigatório 'data_mov' não encontrado no arquivo de movimentos do Protheus.")
        value = row.get(matching_key, "").strip()
        if not value:
            raise DataValidationError(f"Campo 'data_mov' vazio na linha {index}.")
        dates.append(_parse_date(value, field_name="data_mov"))
    return dates

